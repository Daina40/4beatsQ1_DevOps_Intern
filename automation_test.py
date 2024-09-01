from openpyxl import load_workbook
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time

# Load the Excel workbook and select the current day's sheet
workbook = load_workbook("4BeatsQ1.xlsx")
current_day = datetime.now().strftime("%A")  
sheet = workbook[current_day]

# Get the list of keywords from the Excel file
keywords = [cell.value for cell in sheet['C'] if cell.value is not None]

# Set up the Chrome WebDriver
service = Service(executable_path="chromedriver.exe")
driver = webdriver.Chrome(service=service)
driver.maximize_window()
driver.delete_all_cookies()

# Iterate over each keyword and process it
for i, keyword in enumerate(keywords, start=3):
    try:
        driver.get("https://www.google.com")
        search_box = driver.find_element(By.NAME, "q")
        search_box.send_keys(keyword)
        time.sleep(2)
        
        suggestions = driver.find_elements(By.XPATH, "//ul[@role='listbox']//li//span")
        suggestion_texts = [suggestion.text for suggestion in suggestions if suggestion.text.strip()]

        if suggestion_texts:
            longest_option = max(suggestion_texts, key=len)
            shortest_option = min(suggestion_texts, key=len)
            sheet.cell(row=i, column=4, value=longest_option)
            sheet.cell(row=i, column=5, value=shortest_option)

            print(f"Keyword: {keyword} | Longest: {longest_option} | Shortest: {shortest_option}")
        else:
            print(f"No suggestions found for keyword: {keyword}")
    except Exception as e:
        print(f"Error processing keyword: {keyword} | Error: {e}")

# Save the updated Excel file
workbook.save("4BeatsQ1.xlsx")
driver.quit()
